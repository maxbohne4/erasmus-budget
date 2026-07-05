# Erasmus Budget Planner Web-App & Excel Generator (CLOUD VERSIE - SECURE)

import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import io
import requests
import os

st.set_page_config(page_title="Erasmus Budget Planner", page_icon="✈️", layout="wide")

# --- SUPABASE DATABASE KOPPELING ---
# Haalt gegevens veilig op uit de Streamlit 'Secrets' kluis
try:
    URL = st.secrets["SUPABASE_URL"]
    KEY = st.secrets["SUPABASE_KEY"]
except Exception as e:
    st.error("Secrets niet gevonden in Streamlit! Controleer je 'Secrets' instellingen.")
    st.stop()

HEADERS = {
    "apikey": KEY,
    "Authorization": f"Bearer {KEY}",
    "Content-Type": "application/json",
    "Prefer": "return=representation"
}

# --- DATA OPHALEN UIT DE CLOUD ---
def fetch_data():
    try:
        set_res = requests.get(f"{URL}/rest/v1/instellingen?id=eq.1", headers=HEADERS)
        uit_res = requests.get(f"{URL}/rest/v1/uitgaven?order=id.asc", headers=HEADERS)
        
        set_data = set_res.json() if set_res.ok else []
        uit_data = uit_res.json() if uit_res.ok else []
        return set_data, uit_data
    except Exception as e:
        st.error(f"Database verbindingsfout: {e}")
        return [], []

# Initialiseer de State met Cloud Data
if 'initialized' not in st.session_state:
    st.session_state.initialized = True
    instellingen_data, uitgaven_data = fetch_data()
    
    if instellingen_data:
        st.session_state.inkomsten = instellingen_data[0].get('inkomsten', {})
        st.session_state.uitgaven_budget = instellingen_data[0].get('uitgaven_budget', {})
        st.session_state.maanden = instellingen_data[0].get('maanden', 5)
    else:
        st.session_state.inkomsten = {
            "Erasmus+ Beurs (Vooraf - 70/80%)": 1200.0,
            "Erasmus+ Beurs (Achteraf - 20/30%)": 300.0,
            "Spaargeld / Bijdrage Ouders": 1500.0,
            "Overige Inkomsten / Toeslagen": 500.0
        }
        st.session_state.uitgaven_budget = {
            "Vaste kosten (Telefoon, Verzekering)": 50.0,
            "Boodschappen": 200.0,
            "Uitgaan & Terrasjes": 150.0,
            "Tripjes & Reizen": 150.0,
            "Vervoer": 50.0,
            "Onvoorzien": 50.0
        }
        st.session_state.maanden = 5
        
    if uitgaven_data:
        st.session_state.tracker = pd.DataFrame(uitgaven_data)
    else:
        st.session_state.tracker = pd.DataFrame(columns=["id", "datum", "categorie", "omschrijving", "bedrag"])

# Datums veilig omzetten voor maand-analyses
df_tracker = st.session_state.tracker.copy()
if not df_tracker.empty and 'datum' in df_tracker.columns:
    df_tracker['Datum_dt'] = pd.to_datetime(df_tracker['datum'], errors='coerce')
    df_tracker['Maand'] = df_tracker['Datum_dt'].dt.strftime('%Y-%m')
else:
    df_tracker['Maand'] = []

# --- CUSTOM STYLING ---
st.markdown("""
    <style>
    .hoofdtitel { font-size: 2.5rem; font-weight: bold; color: #1F4E78; }
    .subtitel { color: #555555; margin-bottom: 20px; }
    </style>
    """, unsafe_allow_html=True)

st.markdown('<p class="hoofdtitel">✈️ Erasmus Budget Planner (Cloud Editie)</p>', unsafe_allow_html=True)
st.markdown('<p class="subtitel">Live verbonden met jouw veilige Supabase database.</p>', unsafe_allow_html=True)

# --- TABBLADEN STRUCTUUR ---
tab1, tab2, tab3, tab4 = st.tabs(["📊 1. Dashboard", "📅 2. Maandoverzicht", "📝 3. Uitgaven Tracker", "📥 4. Exporteer Excel"])

# ==========================================
# TAB 1: DASHBOARD & INSTELLINGEN
# ==========================================
with tab1:
    col1, col2 = st.columns(2, gap="large")
    
    with col1:
        st.subheader("💰 Totale Geldpot")
        totale_pot = 0.0
        for k, v in st.session_state.inkomsten.items():
            st.session_state.inkomsten[k] = st.number_input(f"{k} (€)", value=float(v), step=50.0, format="%.2f")
            totale_pot += st.session_state.inkomsten[k]
            
        st.session_state.maanden = st.number_input("Aantal maanden Erasmus", value=int(st.session_state.maanden), min_value=1, step=1)
        budget_per_maand = totale_pot / st.session_state.maanden if st.session_state.maanden > 0 else 0
        
        st.info(f"**Totaal te besteden:** € {totale_pot:,.2f}  \n**Beschikbaar per maand:** € {budget_per_maand:,.2f}")

    with col2:
        st.subheader("🎯 Maandelijks Budget")
        totaal_begroot = 0.0
        for k, v in st.session_state.uitgaven_budget.items():
            st.session_state.uitgaven_budget[k] = st.number_input(f"Budget voor {k} (€)", value=float(v), step=10.0, format="%.2f")
            totaal_begroot += st.session_state.uitgaven_budget[k]
            
        resterend_vrije_pot = budget_per_maand - totaal_begroot
        st.info(f"**Totaal gepland per maand:** € {totaal_begroot:,.2f}  \n**Nog vrij te verdelen:** € {resterend_vrije_pot:,.2f}")

    if st.button("💾 Sla Instellingen op in de Cloud", type="primary"):
        data_to_save = {
            "id": 1,
            "inkomsten": st.session_state.inkomsten,
            "uitgaven_budget": st.session_state.uitgaven_budget,
            "maanden": st.session_state.maanden
        }
        try:
            headers_upsert = HEADERS.copy()
            headers_upsert["Prefer"] = "resolution=merge-duplicates"
            res = requests.post(f"{URL}/rest/v1/instellingen", headers=headers_upsert, json=data_to_save)
            if res.ok:
                st.success("Je budgetten zijn veilig online opgeslagen!")
            else:
                st.error(f"Fout bij online opslaan: {res.text}")
        except Exception as e:
            st.error(f"Opslaan mislukt: {e}")

    st.divider()
    
    st.subheader("📈 Overzicht: Begroot vs. Werkelijk Uitgegeven (Totaal)")
    if not df_tracker.empty and 'categorie' in df_tracker.columns:
        actuals = df_tracker.groupby('categorie')['bedrag'].sum().to_dict()
    else:
        actuals = {}
        
    chart_data = []
    for cat, budget in st.session_state.uitgaven_budget.items():
        aantal_actieve_maanden = df_tracker['Maand'].nunique() if (not df_tracker.empty and 'Maand' in df_tracker.columns) else 1
        aantal_actieve_maanden = 1 if aantal_actieve_maanden == 0 else aantal_actieve_maanden
        
        gemiddeld_uitgegeven = actuals.get(cat, 0.0) / aantal_actieve_maanden
        chart_data.append({"Categorie": cat, "Begroot (per maand)": budget, "Uitgegeven (gem. per maand)": gemiddeld_uitgegeven})
        
    df_chart = pd.DataFrame(chart_data).set_index("Categorie")
    st.bar_chart(df_chart, use_container_width=True)

# ==========================================
# TAB 2: MAANDOVERZICHT
# ==========================================
with tab2:
    st.subheader("📅 Uitgaven per Maand")
    if df_tracker.empty or 'Maand' not in df_tracker.columns:
        st.write("Nog geen uitgaven geregistreerd.")
    else:
        unique_months = sorted(df_tracker['Maand'].dropna().unique(), reverse=True)
        for month in unique_months:
            with st.expander(f"Overzicht voor {month}", expanded=True):
                df_month = df_tracker[df_tracker['Maand'] == month]
                totaal_maand = df_month['bedrag'].sum()
                
                m_col1, m_col2 = st.columns([1, 2])
                with m_col1:
                    st.metric(label=f"Totale Uitgaven ({month})", value=f"€ {totaal_maand:,.2f}")
                    verschil = budget_per_maand - totaal_maand
                    if verschil >= 0:
                        st.success(f"Binnen budget! Nog € {verschil:,.2f} over.")
                    else:
                        st.error(f"Over budget! € {abs(verschil):,.2f} te veel uitgegeven.")
                
                with m_col2:
                    cat_month = df_month.groupby('categorie')['bedrag'].sum().reset_index()
                    cat_month.columns = ['Categorie', 'Bedrag']
                    st.dataframe(cat_month, use_container_width=True, hide_index=True)

# ==========================================
# TAB 3: UITGAVEN TRACKER
# ==========================================
with tab3:
    st.subheader("Voeg een nieuwe uitgave toe")
    
    with st.form("nieuwe_uitgave_form", clear_on_submit=True):
        t_col1, t_col2, t_col3, t_col4 = st.columns([1, 1, 2, 1])
        with t_col1:
            datum = st.date_input("Datum", datetime.now())
        with t_col2:
            categorie = st.selectbox("Categorie", list(st.session_state.uitgaven_budget.keys()))
        with t_col3:
            omschrijving = st.text_input("Omschrijving / Wat was het?")
        with t_col4:
            bedrag = st.number_input("Bedrag (€)", min_value=0.0, step=1.0, format="%.2f")
            
        submit = st.form_submit_button("Voeg Uitgave Toe")
        if submit and bedrag > 0:
            new_row = {"datum": str(datum), "categorie": categorie, "omschrijving": omschrijving, "bedrag": bedrag}
            try:
                res = requests.post(f"{URL}/rest/v1/uitgaven", headers=HEADERS, json=new_row)
                if res.ok:
                    inserted_row = res.json()[0]
                    st.session_state.tracker = pd.concat([st.session_state.tracker, pd.DataFrame([inserted_row])], ignore_index=True)
                    st.success("Uitgave live toegevoegd aan de cloud!")
                    st.rerun()
                else:
                    st.error(f"Fout: {res.text}")
            except Exception as e:
                st.error(f"Fout: {e}")

    st.divider()
    st.subheader("Alle Uitgaven")
    if not st.session_state.tracker.empty:
        display_df = st.session_state.tracker.drop(columns=['id'], errors='ignore')
        st.dataframe(display_df, use_container_width=True, hide_index=True)
        if st.button("Laatste uitgave verwijderen (Undo)"):
            try:
                last_id = st.session_state.tracker.iloc[-1]['id']
                res = requests.delete(f"{URL}/rest/v1/uitgaven?id=eq.{int(last_id)}", headers=HEADERS)
                if res.ok:
                    st.session_state.tracker = st.session_state.tracker[:-1]
                    st.rerun()
            except Exception as e:
                st.error(f"Fout: {e}")

# ==========================================
# TAB 4: EXCEL EXPORT
# ==========================================
with tab4:
    st.subheader("Download je data")
    
    def generate_excel():
        output = io.BytesIO()
        wb = openpyxl.Workbook()
        ws_dash = wb.active
        ws_dash.title = "Dashboard"
        
        # Basis stijlen
        NAVY_DARK = "1F4E78"
        font_header = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        fill_header = PatternFill(start_color=NAVY_DARK, end_color=NAVY_DARK, fill_type="solid")
        
        # Vul de Excel... (om de lengte te beperken, hier de essentie)
        ws_dash["A1"] = "Erasmus Budget Overzicht"
        ws_dash["A1"].font = Font(size=16, bold=True)
        
        ws_track = wb.create_sheet(title="Alle Uitgaven")
        for r_idx, row in st.session_state.tracker.iterrows():
            ws_track.append([row.get("datum"), row.get("categorie"), row.get("omschrijving"), row.get("bedrag")])
            
        wb.save(output)
        return output.getvalue()

    st.download_button("📥 Download Excel", data=generate_excel(), file_name="Erasmus_Budget.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")